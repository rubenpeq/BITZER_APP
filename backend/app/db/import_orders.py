# backend/app/db/import_orders.py
"""
Import historical orders/operations/tasks from XLSM files using SQLAlchemy models.

Usage:
  python -m db.import_orders --base-dir /path/to/Orders --preview
  python -m db.import_orders --base-dir /path/to/Orders --db-url "postgresql://..."  # write mode

Options:
  --preview                 : parse & print what would be inserted, do not write DB
  --simulate                : connect to DB and query existence (no writes) to estimate counts
  --update-order-num-pieces : set order.num_pieces = sum of good_pieces (overrides file U2)
"""
import argparse
from pathlib import Path
import re
import datetime
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import os

# Adjust these imports to match your project structure if needed
from db.models import OrderDB, OperationDB, TaskDB

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://bitzer:bitzer123@localhost/orders_db")
FNAME_RE = re.compile(r"(\d+)[_\-]([0-9A-Za-z]{2,8})", re.IGNORECASE)


# ---------- Excel helpers ----------
def excel_cell_to_time(cell_value):
    """
    Normalize a variety of Excel cell time representations into datetime.time or None.

    Supported:
      - Python datetime.datetime / datetime.time -> returns .time()
      - Excel fractional day (0 < x < 1) -> converts to time
      - Excel whole numbers representing HHMM (e.g. 910, 1330) -> interprets as 09:10 / 13:30
      - Strings "910", "09:10", "09:10:00" -> parsed
      - Excel serial dates >= 1 (days since 1899-12-30) -> converted and .time() returned
    """
    import datetime as _dt

    if cell_value is None:
        return None

    # datetime / time
    if isinstance(cell_value, _dt.datetime):
        return cell_value.time()
    if isinstance(cell_value, _dt.time):
        return cell_value

    # numeric values
    if isinstance(cell_value, (int, float)):
        try:
            v = float(cell_value)
        except Exception:
            return None

        # Excel fractional day e.g. 0.375 -> 09:00
        if 0.0 <= v < 1.0:
            seconds = int(round(v * 86400))
            hh = seconds // 3600
            mm = (seconds % 3600) // 60
            ss = seconds % 60
            try:
                return _dt.time(hh, mm, ss)
            except Exception:
                return None

        # Common legacy format: HHMM as integer (e.g. 910 -> 09:10, 1330 -> 13:30)
        # Accept values in [0, 2359] and integer-like floats.
        if 0 <= v < 2400:
            # treat values like 9.1 or 910.0 alike by rounding to nearest integer
            intv = int(round(v))
            hh = intv // 100
            mm = intv % 100
            if 0 <= hh < 24 and 0 <= mm < 60:
                try:
                    return _dt.time(hh, mm, 0)
                except Exception:
                    pass

        # Fallback: if numeric is >= 1 treat as Excel serial date
        try:
            if v >= 1:
                dt = _dt.datetime(1899, 12, 30) + _dt.timedelta(days=v)
                return dt.time()
        except Exception:
            return None

        return None

    # string values
    if isinstance(cell_value, str):
        s = cell_value.strip()
        if s == "":
            return None

        # pure digits: interpret as HHMM if length 3 or 4 or if numeric and within 0-2359
        if re.fullmatch(r"\d{1,4}", s):
            try:
                intv = int(s)
                # if like '9' or '90' ambiguous; treat as HH or HMM only if reasonable
                if 0 <= intv < 2400:
                    hh = intv // 100
                    mm = intv % 100
                    # handle single/2-digit hour like '9' -> 09:00 (intv=9 -> hh=0 mm=9 wrong)
                    # To support '9' as 09:00, treat length
                    if len(s) <= 2 and intv < 24:
                        hh = intv
                        mm = 0
                    if 0 <= hh < 24 and 0 <= mm < 60:
                        return datetime.time(hh, mm, 0)
            except Exception:
                pass

        # common time formats
        for fmt in ("%H:%M:%S", "%H:%M", "%H%M"):
            try:
                return datetime.datetime.strptime(s, fmt).time()
            except Exception:
                pass

        # try replacing separators like '.' or 'h' (e.g. '9.10' or '9h10')
        normalized = re.sub(r"[^\d:]", ":", s)
        parts = normalized.split(":")
        try:
            if len(parts) >= 2:
                hh = int(parts[0])
                mm = int(parts[1])
                if 0 <= hh < 24 and 0 <= mm < 60:
                    return datetime.time(hh, mm, 0)
        except Exception:
            pass

    # unknown format
    return None


def get_order_num_pieces_from_file(path: Path):
    """
    Reads cell U2 (row=2, col=21). Returns int or None.
    """
    try:
        wb = openpyxl.load_workbook(filename=str(path), data_only=True)
        ws = wb.active
        cell = ws.cell(row=2, column=21).value  # U2
        wb.close()
        if cell is None:
            return None
        if isinstance(cell, (int, float)):
            return int(cell)
        s = str(cell).strip()
        if s == "":
            return None
        try:
            return int(float(s))
        except Exception:
            return None
    except Exception:
        return None


def parse_xlsm_rows(path: Path, month_start_date: datetime.date):
    """
    Returns list of task rows (dicts). Reads columns:
      A -> start time
      C -> end time
      F -> num_machines
      G -> num_benches
      J -> pieces -> good_pieces

    This function:
      - parses numeric HHMM (e.g. 910) and string HH:MM formats
      - accepts Excel fractional times (<1) and Excel serial dates (>=1)
      - constructs full datetimes by combining with month_start_date
      - if end_at exists and is before start_at, assumes end_at is next day and adds one day
    """
    wb = openpyxl.load_workbook(filename=str(path), data_only=True)
    ws = wb.active
    rows = []
    # we need up to column 21 (U) for num_pieces reading earlier; tasks use through J (10)
    for row in ws.iter_rows(min_row=1, max_col=21, values_only=True):
        a = row[0]   # A
        c = row[2]   # C
        f = row[5]   # F
        g = row[6]   # G
        j = row[9]   # J

        # skip rows that are entirely empty for the relevant columns
        if a is None and c is None and f is None and g is None and j is None:
            continue

        start_time = excel_cell_to_time(a)
        end_time = excel_cell_to_time(c)

        start_at = None
        end_at = None
        if start_time:
            start_at = datetime.datetime.combine(month_start_date, start_time)
        if end_time:
            end_at = datetime.datetime.combine(month_start_date, end_time)

        # if both present and end is before start, assume end is next day
        if start_at and end_at and end_at < start_at:
            # add one day to end_at
            end_at = end_at + datetime.timedelta(days=1)

        try:
            good_pieces = None if j is None or str(j).strip() == "" else int(float(j))
            bad_pieces = 0 if good_pieces is not None else None
        except Exception:
            good_pieces = None
            bad_pieces = None

        try:
            num_machines = None if f is None or str(f).strip() == "" else int(float(f))
        except Exception:
            num_machines = None
        try:
            num_benches = None if g is None or str(g).strip() == "" else int(float(g))
        except Exception:
            num_benches = None

        rows.append({
            "start_at": start_at,
            "end_at": end_at,
            "num_machines": num_machines,
            "num_benches": num_benches,
            "good_pieces": good_pieces,
            "bad_pieces": bad_pieces,
        })
    wb.close()
    return rows


def infer_month_start_from_folder(folder_name: str):
    m = re.match(r"^\s*(\d{1,2})-(\d{4})\s*$", folder_name)
    if m:
        mm = int(m.group(1)); yy = int(m.group(2))
        return datetime.date(yy, mm, 1)
    return None


# ---------- DB helpers ----------
def make_session(db_url):
    engine = create_engine(db_url)
    Session = sessionmaker(bind=engine)
    return Session()


def ensure_order(session, order_number: int, month_start: datetime.date, num_pieces_from_file=None, preview=False):
    existing = session.query(OrderDB).filter_by(order_number=order_number).one_or_none()
    if existing:
        return existing, False
    if preview:
        return None, True
    new = OrderDB(
        order_number=order_number,
        material_number=0,
        start_date=month_start,
        end_date=None,
        num_pieces=(num_pieces_from_file if num_pieces_from_file is not None else 0),
    )
    session.add(new)
    session.flush()
    return new, True


def ensure_operation(session, order_obj, operation_code: str, preview=False):
    existing = session.query(OperationDB).filter_by(order_id=order_obj.id, operation_code=operation_code).one_or_none()
    if existing:
        return existing, False
    if preview:
        return None, True
    new = OperationDB(order_id=order_obj.id, operation_code=operation_code, machine_id=None)
    session.add(new)
    session.flush()
    return new, True


def insert_task(session, operation_obj, task_payload: dict, preview=False):
    if preview:
        return None
    t = TaskDB(
        operation_id=operation_obj.id,
        process_type=task_payload.get("process_type", "PROCESSING"),
        start_at=task_payload.get("start_at"),
        end_at=task_payload.get("end_at"),
        num_benches=task_payload.get("num_benches"),
        num_machines=task_payload.get("num_machines"),
        good_pieces=task_payload.get("good_pieces"),
        bad_pieces=task_payload.get("bad_pieces"),
        operator_user_id=None,
        operator_bitzer_id=None,
        notes=None
    )
    session.add(t)
    session.flush()
    return t


# ---------- main processing ----------
def process_orders(base_dir: Path, db_url: str, preview: bool, simulate: bool, update_order_num_pieces: bool):
    session = None
    if not preview and not simulate:
        session = make_session(db_url)
    elif simulate:
        session = make_session(db_url)

    stats = {"orders_created": 0, "orders_existing": 0, "operations_created": 0, "operations_existing": 0, "tasks_inserted": 0}
    files = sorted(base_dir.rglob("*.xlsm"))
    order_good_pieces = {}
    file_order_numpieces = {}

    for f in files:
        parent = f.parent.name
        month_start = infer_month_start_from_folder(parent)
        if month_start is None:
            month_start = infer_month_start_from_folder(f.parent.parent.name) or datetime.date.today().replace(day=1)

        m = FNAME_RE.search(f.name)
        if not m:
            print("Skipping file (name not matched):", f)
            continue
        order_number = int(m.group(1))
        operation_code = m.group(2)

        # read possible order.num_pieces from U2
        num_pieces_from_file = get_order_num_pieces_from_file(f)
        # parse rows
        rows = parse_xlsm_rows(f, month_start)

        if preview and not simulate:
            print("FILE:", f, "month_start:", month_start, "order:", order_number, "op:", operation_code, "rows:", len(rows), "order.num_pieces(U2):", num_pieces_from_file)
            for r in rows[:5]:
                print("   ", r)
            stats["tasks_inserted"] += len(rows)
            for r in rows:
                if r.get("good_pieces"):
                    order_good_pieces.setdefault(order_number, 0)
                    order_good_pieces[order_number] += int(r["good_pieces"])
            if num_pieces_from_file is not None:
                file_order_numpieces.setdefault(order_number, num_pieces_from_file)
            continue

        if simulate and preview:
            print("SIMULATE FILE:", f, "month_start:", month_start, "order:", order_number, "op:", operation_code, "rows:", len(rows), "order.num_pieces(U2):", num_pieces_from_file)
            order_obj = session.query(OrderDB).filter_by(order_number=order_number).one_or_none()
            if order_obj:
                stats["orders_existing"] += 1
            else:
                stats["orders_created"] += 1
            if order_obj:
                op_obj = session.query(OperationDB).filter_by(order_id=order_obj.id, operation_code=operation_code).one_or_none()
                if op_obj:
                    stats["operations_existing"] += 1
                else:
                    stats["operations_created"] += 1
            else:
                stats["operations_created"] += 1
            stats["tasks_inserted"] += len(rows)
            for r in rows:
                if r.get("good_pieces"):
                    order_good_pieces.setdefault(order_number, 0)
                    order_good_pieces[order_number] += int(r["good_pieces"])
            if num_pieces_from_file is not None:
                file_order_numpieces.setdefault(order_number, num_pieces_from_file)
            continue

        # write mode
        try:
            order_obj, created_order = ensure_order(session, order_number, month_start, num_pieces_from_file, preview=False)
            if created_order:
                stats["orders_created"] += 1
                print("Created order", order_number, "num_pieces(from file):", num_pieces_from_file)
            else:
                stats["orders_existing"] += 1

            op_obj, created_op = ensure_operation(session, order_obj, operation_code, preview=False)
            if created_op:
                stats["operations_created"] += 1
                print("Created operation", operation_code, "for order", order_number)
            else:
                stats["operations_existing"] += 1

            for r in rows:
                payload = dict(r)
                payload["process_type"] = "PROCESSING"
                insert_task(session, op_obj, payload, preview=False)
                stats["tasks_inserted"] += 1
                if payload.get("good_pieces"):
                    order_good_pieces.setdefault(order_number, 0)
                    order_good_pieces[order_number] += int(payload["good_pieces"])

            session.commit()
            if num_pieces_from_file is not None:
                # remember file-provided num_pieces for later reporting (won't override if --update-order-num-pieces)
                file_order_numpieces.setdefault(order_number, num_pieces_from_file)
        except Exception as e:
            session.rollback()
            print("ERROR while processing", f, "->", e)

    # optionally update Order.num_pieces sums after import
    if not preview and update_order_num_pieces and session:
        for order_number, total_good in order_good_pieces.items():
            try:
                obj = session.query(OrderDB).filter_by(order_number=order_number).one_or_none()
                if obj:
                    obj.num_pieces = total_good
                    print(f"Set order {order_number}.num_pieces = {total_good}")
            except Exception as e:
                print("Error updating order num_pieces for", order_number, e)
        session.commit()

    # final report (include any file-provided num_pieces seen)
    if file_order_numpieces:
        print("File-provided order.num_pieces (U2) samples:")
        for k, v in list(file_order_numpieces.items())[:20]:
            print(" ", k, "->", v)

    if session:
        session.close()
    return stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base-dir", required=True, help="Base Orders directory (with subfolders like 01-2020)")
    ap.add_argument("--db-url", required=False, help="Postgres DSN (overrides env)")
    ap.add_argument("--preview", action="store_true", help="Preview only (no DB writes)")
    ap.add_argument("--simulate", action="store_true", help="Query DB and report counts of would-be-created objects (no writes)")
    ap.add_argument("--update-order-num-pieces", action="store_true", help="After import, set order.num_pieces = sum of good_pieces")
    args = ap.parse_args()

    base = Path(args.base_dir)
    if not base.exists():
        print("Base dir doesn't exist:", base)
        return

    db_url = args.db_url or DATABASE_URL
    stats = process_orders(base, db_url, args.preview, args.simulate, args.update_order_num_pieces)
    print("DONE. stats:", stats)


if __name__ == "__main__":
    main()
